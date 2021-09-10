from openpyxl import Workbook, load_workbook
import time

class excel():

    def __init__(self) -> None:
        self.wb = Workbook()
        self.wbac = self.wb.active

    def get_sheet_names(self):
        return self.wb.get_sheet_names()
    # 向sheet中写入一行数据
    def write_row(self, list, sheet_name='Sheet', with_url=None):
        # sheets = self.get_sheet_names()
        sheet = self.wb.get_sheet_by_name(sheet_name)
        #sheet = wb.get_sheet_by_name(sheets[0])
        
        if sheet:
            # row = [y for item in list for y in item]
            # sheet.append(row)
            for i in range(len(list)):
                # ll = []
                row = list[i]
                for y in range(len(row)):
                    content = row[y]
                    if with_url == y and i > 0:
                        # print(i,y)
                        # 替换成链接
                        # wu = '=HYPERLINK("{}", "{}")'.format(row[y+1], content)
                        list[i][y] = '=HYPERLINK("{}", "{}")'.format(row[y+1], content)
                        # v = self.wbac.cell(row=i+1, column=y+1).value = '=HYPERLINK("{}", "{}")'.format(row[y+1], content)
                #         ll.append(wu)
                #         continue
                #     else:
                #         ll.append(content)
                # print(ll)
                # sheet.append(ll)
                sheet.append(row)



    # def write_row_url(self, list, sheet_name='Sheet'):
    #     # sheets = self.get_sheet_names()
    #     sheet = self.wb.get_sheet_by_name(sheet_name)
    #     #sheet = wb.get_sheet_by_name(sheets[0])
    #     if sheet:
    #         row = [y for item in list for y in item]
    #         print(row)
    #         sheet.append(row)

    def new_sheet(self, name):
        self.wb.create_sheet(name)

    def save(self, path=f"./files/Excel{str(round(time.time()*1000))}.xlsx"):
        self.wb.save(path)

    # def __call__(self, list):
    #     self
    #     pass

    
# 新建excel，并创建多个sheet
if __name__ == "__main__":

    ex = excel()

    ex.new_sheet('啊啊啊')
    names = ex.get_sheet_names()
    print(names)

    ats = [['创作','标题','链接','描述','时间','阅读','评论']]
    ats.append(['原创', '整个大活，采集8个代理IP站点，为Python代理池铺路，爬虫120例之第15例', 'https://dream.blog.csdn.net/article/details/119137580', 'Python爬虫代理池，入门级的，来看看 吧', '2021-07-28 10:05:36', '3705', '23'])
    ats.append(['原创', '国内首创滚雪球学编程之 Python 技术栈文章清单（还在更新中）', 'https://dream.blog.csdn.net/article/details/114586600', '滚雪球学 Python 系列专栏，国产区一套非常牛 的课程。', '2021-03-09 16:24:09', '8376', '34'])
    ex.write_row(ats, with_url=1)
    ex.save()

    # # default = book.get_sheet_by_name(names[0])
    # # book.remove(default)
    # # 2 新建自定义的sheet
    # # for i in range(1, 3):
    #     # 为每个sheet设置title，插入位置index
    # #    new_sheet("Sheet" + str(i))
    #     # sheet = book.create_sheet("Sheet" + str(i), i)
    # sheets = book.get_sheet_names()
    # print(sheets)
    # # count = 0
    # # 4 向sheet中插入数据
    # sheet = book.get_sheet_by_name(sheets[0])

    # for i in range(0, 3):
    #     row = [i for i in range(100)]
    #     #sheet.append(row)
    #     if sheet:
    #         write_row(row, sheet)

    #     # insertOne("ni", book.get_sheet_by_name(sheets[1]))
    #     # insertOne("wo", book.get_sheet_by_name(sheets[0]))
    #     # insertOne("ta", book.get_sheet_by_name(sheets[1]))
    #     # count = count + 1

    # # 5 保存数据到.xlsx文件
    # save()
    # # print(str(count))