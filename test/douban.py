import requests, random, time, pymysql
from bs4 import BeautifulSoup
from openpyxl import Workbook

# 获取列表
def get_books(html, books):
    soup = BeautifulSoup(html, 'html.parser')
    items = soup.select('table')

    print(f'正在爬取{len(items)}个')
    if items:
        for i in range(len(items)):
            book = []
            item = items[i]
            # 文字内容
            div = item.find(name='div', attrs={'class': 'pl2'})
            a = div.find('a')
            title = a['title']
            book.append(title)

            eng_title = div.find('span')
            if eng_title:
                book.append(eng_title.string)
            else: 
                book.append('')

            # 属性？
            attr = item.find(name='p', attrs={'class': 'pl'})
            if attr:
                info = attr.string.split("/")
                # 有些书是外国的，会有翻译者，而有些书是国内的，没有翻译者，如果用正数下标去找翻译者后面的属性，就不好判断，直接通过复数下标刚好可以跳过翻译者
                book.append(info[0].strip())
                book.append(info[-3].strip())
                book.append(info[-2].strip())
                book.append(info[-1].replace('元', ''))
            else: 
                book.append('')

            # 评分
            stars_div = item.find(name='div', attrs={'class': 'star'})
            star = stars_div.find(name='span', attrs={'class': 'rating_nums'}).string
            count = stars_div.find(name='span', attrs={'class': 'pl'}).string.replace('\n', '').replace(' ', '').replace('(', '').replace(')', '').replace('人评价', '')
            
            book.append(star)
            book.append(count)

            # 语录
            quote = item.find(name='p', attrs={'class': 'quote'})
            if quote:
                inq = quote.find(name='span', attrs={'class': 'inq'})
                book.append(inq.string)
            else: 
                book.append('')
            
            books.append(book)
    
    return books

# 保存cvs
def wrt(books):
    name = str(round(time.time()))
    with open(f'./files/{name}.csv', 'a+', encoding='utf-8') as f:
        for i in range(len(books)):
            f.write(','.join(books[i]))
            f.write('\n')
        #print(','.join(i))

# 获取html
def get_html(url, ret_type="text", timeout=50, encoding="utf-8"):
	headers = get_headers()
	res = requests.get(url, headers=headers, timeout=timeout)
	res.encoding = encoding
	if ret_type == "text":
		return res.text
	elif ret_type == "image":
		return res.content
	elif ret_type == "json":
		return res.json()

# 获取请求头
def get_headers(localhost=True, refer="https://www.baidu.com", host=None):
	ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"
	if not localhost:
		uas = [
			"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36",
			"Mozilla/5.0 (compatible; Baiduspider/2.0; +http://www.baidu.com/search/spider.html)",
			"Mozilla/5.0 (compatible; Baiduspider-render/2.0; +http://www.baidu.com/search/spider.html)",
			"Baiduspider-image+(+http://www.baidu.com/search/spider.htm)",
			"Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",
			"Mozilla/5.0 (compatible; Googlebot-Image/1.0; +http://www.google.com/bot.html)",
			"Sogou web spider/4.0(+http://www.sogou.com/docs/help/webmasters.htm#07)",
			"Sogou News Spider/4.0(+http://www.sogou.com/docs/help/webmasters.htm#07)",
			"Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0);",
			"Mozilla/5.0 (compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm)",
			"Sosospider+(+http://help.soso.com/webspider.htm)",
			"Mozilla/5.0 (compatible; Yahoo! Slurp China; http://misc.yahoo.com.cn/help.html)"
		]
		ua = random.choice(uas)
	headers = {
		"User-Agent": ua,
		"Referer": refer,
		"Host": host
	}
	return headers

# 自定义个excel，可以将url给到文本上
class excel():

    def __init__(self) -> None:
        self.wb = Workbook()
        self.wbac = self.wb.active

    def get_sheet_names(self):
        return self.wb.get_sheet_names()
    # 向sheet中写入一行数据
    def write_row(self, list, sheet_name='Sheet', with_url=None):
        # sheets = self.get_sheet_names()
        sheet = self.wb[sheet_name]
        #sheet = wb.get_sheet_by_name(sheets[0])
        
        if sheet:
            # row = [y for item in list for y in item]
            # sheet.append(row)
            for i in range(len(list)):
                # ll = []
                row = list[i]
                for y in range(len(row)):
                    content = row[y]
                    if with_url == y and i > 0:
                        # print(i,y)
                        # 替换成链接
                        # wu = '=HYPERLINK("{}", "{}")'.format(row[y+1], content)
                        list[i][y] = '=HYPERLINK("{}", "{}")'.format(row[y+1], content)
                        # v = self.wbac.cell(row=i+1, column=y+1).value = '=HYPERLINK("{}", "{}")'.format(row[y+1], content)
                #         ll.append(wu)
                #         continue
                #     else:
                #         ll.append(content)
                # print(ll)
                # sheet.append(ll)
                sheet.append(row)

    def new_sheet(self, name):
        self.wb.create_sheet(name)

    def save(self, path=f"./files/Excel{str(round(time.time()*1000))}.xlsx"):
        self.wb.save(path)

    # def __call__(self, list):
    #     self
    #     pass

# 获取数据库连接
def get_connection(user, pwd, host='localhost', database='sqc'):
    # 打开数据库连接
    global db
    db = pymysql.connect(host=host, user=user, password=pwd, database=database)
    get_cursor()
    return db

def get_cursor():
    global cursor
    cursor = db.cursor()

# 更新
def upd(sql):
    print(sql)
    try:
        # 执行SQL语句
        cursor.execute(sql)
    # 提交到数据库执行
        db.commit()
    except:
        # 发生错误时回滚
        db.rollback()

# 保存数据库
def savedb(books):
    
    db = get_connection('root', 'root')

    for i in range(1, len(books)):
        book = books[i]
        # 使用 execute()  方法执行 SQL 查询
        # url = f'{url}?start={i*limit}'

        book_name = book[0]
        book_other_name = book[1]
        author = book[2]
        publisher = book[3]
        publisher_time = book[4]
        price = book[5]

        sql = f"Insert into books (`book_name`,`book_other_name`,`author`,`publisher`,`publisher_time`,`price`) values ('{book_name}','{book_other_name}','{author}','{publisher}','{publisher_time}',{price})"
        
        upd(sql)

        
    # # 使用 fetchone() 方法获取单条数据.
    # data = cursor.fetchall()

    # print(f"result : {data}")

    # 关闭数据库连接
    db.close()

if __name__ == '__main__':

    url = 'https://book.douban.com/top250?start=25'
    # https://book.douban.com/top250?start=0
    # https://book.douban.com/top250?start=25
    # https://book.douban.com/top250?start=50
    # 分析url发现，每页25

    limit = 25
    # 要爬取的页数
    pages = 2

    books = [['书名','英文名','作者','出版社','发布时间','价格','评分','人数','语录']]
    # 循环获取每一页
    for i in range(0, pages):
        url = f'{url}?start={i*limit}'
        print(f'爬取。。。{url}')
        html = get_html(url)

        # print(html)
        books = get_books(html, books)
        time.sleep(3) # 3秒
    # 保存csv，保存后如若用excel打开，则需先用记事本打开，另存为 ANSI ， 才不会乱码
    #wrt(books)

    if len(books) > 0: 
        # 保存excel
        ex = excel()
        # with_url , 哪一列下标内容带url，ats数组改列后面必须紧接着url
        ex.write_row(books)
        #ex.wbac.cell(row=1, column=1).value = '=HYPERLINK("{}", "{}")'.format('https://www.baidu.com', "Link Name")
        ex.save()

        # 存数据库
        # books = [['书名','英文名','作者','出版社','发布时间','价格','评分','人数','语录'],['红楼梦','','[清] 曹雪芹 著','人民文学出版社','1996-12','59.70'],['红楼梦1','','[清] 曹雪芹 著','人民文学出版社','1996-12','59.70']]
        savedb(books)
